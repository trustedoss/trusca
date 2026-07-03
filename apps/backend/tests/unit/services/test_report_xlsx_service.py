"""
Unit tests for the XLSX vulnerability-report builder — Phase G (P3-8).

``build_report_xlsx`` is pure (no I/O / no DB), so these open the produced bytes
with openpyxl and assert the workbook shape, the data-row values, and — most
importantly — the formula-injection defence: a component / summary whose text
starts with ``= + - @`` must be neutralised with a leading ``'`` so a hostile
package name can't execute as a spreadsheet formula (CWE-1236).
"""

from __future__ import annotations

from datetime import UTC, date, datetime
from io import BytesIO

from openpyxl import load_workbook

from services.report_xlsx_service import XLSX_MIME, build_report_xlsx

_GENERATED = datetime(2026, 7, 3, 12, 0, tzinfo=UTC)


def _build(**overrides):
    kwargs = dict(
        project_name="acme/checkout",
        generated_at=_GENERATED,
        risk_score=72.5,
        total_components=2,
        severity_distribution={"critical": 1, "high": 2, "low": 0},
        license_distribution={"allowed": 2, "forbidden": 1},
        components=[
            {
                "name": "lodash",
                "version": "4.17.19",
                "purl": "pkg:npm/lodash@4.17.19",
                "license": "MIT",
                "severity_max": "high",
                "vulnerability_count": 2,
                "direct": True,
            },
        ],
        vulnerabilities=[
            {
                "cve_id": "CVE-2099-0001",
                "severity": "critical",
                "cvss_score": 9.8,
                "epss_score": 0.5,
                "epss_percentile": 0.9,
                "kev": True,
                "kev_due_date": date(2026, 7, 20),
                "status": "new",
                "component_name": "lodash",
                "component_version": "4.17.19",
                "summary": "prototype pollution",
            },
        ],
        vulnerabilities_total=1,
    )
    kwargs.update(overrides)
    return build_report_xlsx(**kwargs)


def _load(data: bytes):
    assert data[:2] == b"PK", "not a valid xlsx (zip) container"
    return load_workbook(BytesIO(data))


def test_workbook_has_the_three_named_sheets() -> None:
    wb = _load(_build())
    assert wb.sheetnames == ["Overview", "Components", "Vulnerabilities"]


def test_mime_is_the_openxml_spreadsheet_type() -> None:
    assert XLSX_MIME == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_overview_sheet_carries_project_and_totals() -> None:
    wb = _load(_build())
    ov = wb["Overview"]
    # (field, value) pairs — collect into a dict for order-independent asserts.
    kv = {row[0].value: row[1].value for row in ov.iter_rows() if row[0].value}
    assert kv["Project"] == "acme/checkout"
    assert kv["Generated (UTC)"] == _GENERATED.isoformat()
    assert kv["Risk score"] == 72.5
    assert kv["Total components"] == 2
    assert kv["Total vulnerabilities"] == 1


def test_components_sheet_header_and_row() -> None:
    wb = _load(_build())
    ws = wb["Components"]
    header = [c.value for c in ws[1]]
    assert header == [
        "Name",
        "Version",
        "PURL",
        "License",
        "Max severity",
        "Vulnerabilities",
        "Direct",
    ]
    row = [c.value for c in ws[2]]
    assert row[0] == "lodash"
    assert row[1] == "4.17.19"
    assert row[5] == 2  # vulnerability_count is a native int, not text
    assert row[6] is True  # direct is a native bool


def test_vulnerabilities_sheet_header_and_kev_columns() -> None:
    wb = _load(_build())
    ws = wb["Vulnerabilities"]
    header = [c.value for c in ws[1]]
    assert header == [
        "CVE",
        "Severity",
        "CVSS",
        "EPSS",
        "EPSS percentile",
        "KEV",
        "KEV due date",
        "Status",
        "Component",
        "Version",
        "Summary",
    ]
    row = [c.value for c in ws[2]]
    assert row[0] == "CVE-2099-0001"
    assert row[5] is True  # KEV native bool
    assert row[6] == "2026-07-20"  # date → ISO string, no Excel serial


def test_formula_injection_is_neutralised_in_string_cells() -> None:
    """A hostile component name / summary must not stay a live formula."""
    wb = _load(
        _build(
            components=[
                {
                    "name": "=cmd|'/c calc'!A0",
                    "version": "-1+1",
                    "purl": "@SUM(A1)",
                    "license": "MIT",
                    "severity_max": "high",
                    "vulnerability_count": 0,
                    "direct": False,
                },
            ],
            vulnerabilities=[
                {
                    "cve_id": "CVE-2099-0002",
                    "severity": "high",
                    "cvss_score": 7.5,
                    "epss_score": None,
                    "epss_percentile": None,
                    "kev": False,
                    "kev_due_date": None,
                    "status": "new",
                    "component_name": "=HYPERLINK(1)",
                    "component_version": "1.0",
                    "summary": "+SUM(A1)",
                },
            ],
        )
    )
    comp = wb["Components"]
    assert comp.cell(row=2, column=1).value == "'=cmd|'/c calc'!A0"
    assert comp.cell(row=2, column=2).value == "'-1+1"
    assert comp.cell(row=2, column=3).value == "'@SUM(A1)"
    vuln = wb["Vulnerabilities"]
    assert vuln.cell(row=2, column=9).value == "'=HYPERLINK(1)"
    assert vuln.cell(row=2, column=11).value == "'+SUM(A1)"
    # None cells render empty, never a leading-quote artefact.
    assert vuln.cell(row=2, column=4).value in (None, "")


def test_pure_function_is_deterministic() -> None:
    """Identical inputs → identical bytes (no timestamps baked by us)."""
    assert _build() == _build()


def test_empty_project_still_builds_valid_workbook() -> None:
    wb = _load(_build(components=[], vulnerabilities=[], vulnerabilities_total=0))
    assert wb.sheetnames == ["Overview", "Components", "Vulnerabilities"]
    # Header rows present even with no data rows.
    assert wb["Components"].max_row == 1
    assert wb["Vulnerabilities"].max_row == 1
