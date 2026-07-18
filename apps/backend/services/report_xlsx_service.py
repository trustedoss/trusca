"""
XLSX vulnerability-report builder — Phase G (P3-8, the CLAUDE.md "Excel / PDF
reports" promise).

``build_report_xlsx`` is the spreadsheet sibling of
:func:`services.report_service.build_report_html`: a **pure** function that turns
the exact dict shapes the report read-services already return
(``get_project_overview`` / ``list_components_for_project`` /
``list_project_vulnerabilities``) into an ``.xlsx`` workbook as bytes. It performs
no I/O and no DB access — identical inputs produce identical bytes — so the
endpoint offloads it to a threadpool exactly like the weasyprint PDF path.

Three worksheets:
  - **Overview** — project, generated-at, risk score, component total, plus the
    severity- and license-distribution breakdowns.
  - **Components** — one row per component (name, version, purl, license, max
    severity, vulnerability count, direct/transitive).
  - **Vulnerabilities** — one row per finding (CVE, severity, CVSS, EPSS, KEV +
    due date, status, affected component + version, summary).

Formula-injection defence (CWE-1236 / OWASP CSV-injection)
----------------------------------------------------------
Component names, purls, CVE ids, licenses and summaries all originate from
scanned third-party metadata, so a hostile package could name itself
``=cmd|'/c calc'!A0``. Excel / LibreOffice / Sheets execute a cell whose text
starts with ``= + - @`` (or a leading tab / CR) as a formula when the file is
opened. Every STRING cell is therefore routed through :func:`_safe`, which
prepends a single quote to a value beginning with one of those characters —
mirroring ``admin_audit_service._csv_cell`` (the audit-log CSV export's defence).
Numbers / booleans are written as native types and can never be a formula.
"""

from __future__ import annotations

import re
import zipfile
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

# openpyxl rewrites ``docProps/core.xml``'s ``<dcterms:modified>`` with
# ``datetime.now()`` at save time regardless of ``wb.properties.modified`` — so
# it must be re-pinned in the produced bytes (see ``_normalize_zip_timestamps``).
_CORE_XML = "docProps/core.xml"
_CORE_DT_RE = re.compile(
    r"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</dcterms:(?:created|modified)>)"
)

# Leading characters that make a spreadsheet cell execute as a formula. Same set
# as ``admin_audit_service._DANGEROUS_CSV_PREFIX`` (OWASP CSV-injection).
_DANGEROUS_PREFIX = ("=", "+", "-", "@", "\t", "\r")

# Fixed severity order for the Overview breakdown (matches the PDF report).
_SEVERITY_ORDER = ("critical", "high", "medium", "low", "info", "none", "unknown")

_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _safe(value: Any) -> Any:
    """Return a spreadsheet-safe cell value.

    ``None`` → empty string. ``datetime`` / ``date`` → ISO-8601 text (so the
    cell never carries a locale-dependent Excel serial). Numbers and booleans
    pass through as native types. Any other value is stringified, and a string
    that begins with a formula-trigger character is prefixed with ``'`` so the
    spreadsheet renders it as literal text, never a formula.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, int | float):
        return value
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    rendered = str(value)
    if rendered and rendered[0] in _DANGEROUS_PREFIX:
        return "'" + rendered
    return rendered


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold


def _autosize(ws: Worksheet, widths: dict[int, int]) -> None:
    """Set a fixed width per 1-based column index (simple, no measurement)."""
    for idx, width in widths.items():
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width


def _build_overview_sheet(
    ws: Worksheet,
    *,
    project_name: str,
    generated_at: datetime,
    risk_score: float,
    total_components: int,
    total_vulnerabilities: int,
    severity_distribution: dict[str, int],
    license_distribution: dict[str, int],
) -> None:
    ws.title = "Overview"
    ws.append(["Field", "Value"])
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    ws.append([_safe("Project"), _safe(project_name)])
    ws.append([_safe("Generated (UTC)"), _safe(generated_at)])
    ws.append([_safe("Risk score"), _safe(risk_score)])
    ws.append([_safe("Total components"), _safe(total_components)])
    ws.append([_safe("Total vulnerabilities"), _safe(total_vulnerabilities)])

    ws.append([])
    section = ws.cell(row=ws.max_row + 1, column=1, value="Severity distribution")
    section.font = bold
    for key in _SEVERITY_ORDER:
        if key in severity_distribution:
            ws.append([_safe(key), _safe(int(severity_distribution.get(key, 0)))])

    ws.append([])
    section = ws.cell(row=ws.max_row + 1, column=1, value="License distribution")
    section.font = bold
    # Deterministic order regardless of dict insertion order.
    for key in sorted(license_distribution):
        ws.append([_safe(key), _safe(int(license_distribution.get(key, 0)))])

    _autosize(ws, {1: 24, 2: 40})


def _build_components_sheet(ws: Worksheet, components: list[dict[str, Any]]) -> None:
    ws.title = "Components"
    _write_header(
        ws,
        ["Name", "Version", "PURL", "License", "Max severity", "Vulnerabilities", "Direct"],
    )
    for c in components:
        ws.append(
            [
                _safe(c.get("name")),
                _safe(c.get("version")),
                _safe(c.get("purl")),
                _safe(c.get("license")),
                _safe(c.get("severity_max")),
                _safe(int(c.get("vulnerability_count") or 0)),
                _safe(bool(c.get("direct"))),
            ]
        )
    _autosize(ws, {1: 30, 2: 14, 3: 48, 4: 20, 5: 14, 6: 16, 7: 8})


def _build_vulnerabilities_sheet(
    ws: Worksheet, vulnerabilities: list[dict[str, Any]]
) -> None:
    ws.title = "Vulnerabilities"
    _write_header(
        ws,
        [
            "CVE",
            "Severity",
            "CVSS",
            "EPSS",
            "EPSS percentile",
            "KEV",
            "KEV due date",
            "Status",
            "Component",
            "Version",
            "Summary",
        ],
    )
    for v in vulnerabilities:
        ws.append(
            [
                _safe(v.get("cve_id")),
                _safe(v.get("severity")),
                _safe(v.get("cvss_score")),
                _safe(v.get("epss_score")),
                _safe(v.get("epss_percentile")),
                _safe(bool(v.get("kev"))),
                _safe(v.get("kev_due_date")),
                _safe(v.get("status")),
                _safe(v.get("component_name")),
                _safe(v.get("component_version")),
                _safe(v.get("summary")),
            ]
        )
    _autosize(
        ws,
        {1: 18, 2: 10, 3: 8, 4: 10, 5: 16, 6: 6, 7: 14, 8: 14, 9: 30, 10: 14, 11: 60},
    )


def build_report_xlsx(
    *,
    project_name: str,
    generated_at: datetime,
    risk_score: float,
    total_components: int,
    severity_distribution: dict[str, int],
    license_distribution: dict[str, int],
    components: list[dict[str, Any]],
    vulnerabilities: list[dict[str, Any]],
    components_total: int | None = None,
    vulnerabilities_total: int | None = None,
) -> bytes:
    """Build the vulnerability report as an ``.xlsx`` workbook (bytes).

    Pure: no I/O, no DB. ``components`` / ``vulnerabilities`` are the dict shapes
    returned by ``list_components_for_project`` / ``list_project_vulnerabilities``
    (read defensively with ``.get`` so a future column never breaks the export).
    The signature mirrors :func:`services.report_service.build_report_html` so
    the endpoint feeds both from one set of read-service calls.
    """
    wb = Workbook()
    # Pin the document metadata to the report's ``generated_at`` (and a fixed
    # creator) so the function stays pure: openpyxl otherwise stamps
    # ``core.xml`` with ``datetime.now()`` at Workbook() time, which would make
    # two builds of the same inputs differ byte-for-byte and leak the build
    # host's clock. ``generated_at`` is already an input, so this keeps
    # "same inputs → same bytes".
    stamp = generated_at.replace(tzinfo=None)
    wb.properties.created = stamp
    wb.properties.modified = stamp
    wb.properties.creator = "TrustedOSS Portal"
    wb.properties.lastModifiedBy = "TrustedOSS Portal"
    overview_ws = wb.active
    _build_overview_sheet(
        overview_ws,
        project_name=project_name,
        generated_at=generated_at,
        risk_score=risk_score,
        total_components=total_components,
        total_vulnerabilities=(
            vulnerabilities_total
            if vulnerabilities_total is not None
            else len(vulnerabilities)
        ),
        severity_distribution=severity_distribution,
        license_distribution=license_distribution,
    )
    _build_components_sheet(wb.create_sheet("Components"), components)
    _build_vulnerabilities_sheet(wb.create_sheet("Vulnerabilities"), vulnerabilities)

    buffer = BytesIO()
    wb.save(buffer)
    return _normalize_zip_timestamps(buffer.getvalue(), stamp)


def _normalize_zip_timestamps(data: bytes, stamp: datetime) -> bytes:
    """Rewrite an xlsx (a zip) with fixed member timestamps → byte-determinism.

    Two wall-clock leaks make the "pure" builder non-deterministic across a
    one-second boundary:

    1. ``wb.save`` stamps each zip member's ``date_time`` with the clock at save
       time (openpyxl offers no hook to pin it).
    2. openpyxl overwrites ``docProps/core.xml``'s ``<dcterms:modified>`` with
       ``datetime.now()`` at save, ignoring ``wb.properties.modified``.

    Both are re-pinned to ``stamp`` (already an input, so "same inputs → same
    bytes" holds), preserving member order, contents, compression type, and
    attributes. The zip's earliest representable time is 1980-01-01;
    ``generated_at`` is always well after that.
    """
    date_time = (stamp.year, stamp.month, stamp.day, stamp.hour, stamp.minute, stamp.second)
    iso = stamp.strftime("%Y-%m-%dT%H:%M:%SZ")
    out = BytesIO()
    with zipfile.ZipFile(BytesIO(data), "r") as src, zipfile.ZipFile(out, "w") as dst:
        for info in src.infolist():
            member = zipfile.ZipInfo(filename=info.filename, date_time=date_time)
            member.compress_type = info.compress_type
            member.external_attr = info.external_attr
            member.internal_attr = info.internal_attr
            member.create_system = info.create_system
            content = src.read(info.filename)
            if info.filename == _CORE_XML:
                text = content.decode("utf-8")
                text = _CORE_DT_RE.sub(rf"\g<1>{iso}\g<2>", text)
                content = text.encode("utf-8")
            dst.writestr(member, content)
    return out.getvalue()


__all__ = ["build_report_xlsx", "XLSX_MIME"]

XLSX_MIME = _MIME
